"""Утилиты для работы с файлами"""
from xlsxwriter import Workbook
from config import LIST_WITH_COLUMNS_COMBINED
from openpyxl import load_workbook
import logging # Импортируем библиотеку для безопасного хранения логов


logger = logging.getLogger(__name__)

list_with_filtered_values = []  # список с отфильтрованными значениям (куки)


def write_to_excel(list_with_data, keyword='combined'):
    """Производит запись в xlsx файл"""
    wb = Workbook(f'{keyword}.xlsx')
    ws = wb.add_worksheet(f'{keyword}') # По дефоолту 'sheet 1'
    ordered_list = LIST_WITH_COLUMNS_COMBINED
    first_row = 0
    for header in ordered_list:
        col = ordered_list.index(header) # Порядок столбцов
        ws.write(first_row, col, header) # Первая строка - заголовок

    row = 1
    for dict_ in list_with_data:
        for _key,_value in dict_.items():
            col=ordered_list.index(_key)
            ws.write(row,col,_value)
        row += 1 # Перевод на следующую строку
    wb.close()


def read_txt_file_and_lines_to_list(file_name='combined.txt'):
    """Читает файл txt построчно, РАЗДЕЛЯЕТ СТРОКИ по ":" и возвращает результат в виде списка со списками .
    """
    with open(file_name) as file:
        user_data_list = []
        while line := file.readline():
            cur_line = line.rstrip()
            user_data = cur_line.split(':')
            user_data_list.append(user_data)
        return user_data_list


def read_txt_file_return_list_with_lines_text(file_name='messages.txt'):
    """Читает файл txt построчно и возвращает результат в виде списка.
    """
    with open(file_name, encoding='utf-8', errors='ignore') as file:
        list_with_lines_text = []
        while line := file.readline():
            list_with_lines_text.append(line.strip('\n'))
        return list_with_lines_text

def open_xlsx_file_and_return_active_sheet(file_name='combined.xlsx'):
    """Открывает файл xlsx и возвращает книгу и текущий лист.
    """
    # Найти целевой файл и рабочий лист
    try:
        workbook = load_workbook(file_name) 
        worksheet = workbook.active  # делаем единственный лист активным 
        if worksheet:
            logger.debug(f'Рабочий лист {worksheet} найден')
        else:
            logger.debug(f'Рабочий лист {worksheet} не найден')
    except:
        logger.critical('КРИТИЧЕСКАЯ ОШИБКА: xlsx файл или лист. Получение данных невозможно.')
    return workbook, worksheet

    
def read_parameters_from_txt_file_and_add_to_dict(file_name='config.txt'):
    """Читает файл txt построчно и возвращает результат в виде словаря.
    """
    dict_parameter_value = {}
    with open(file_name) as file:
        while line := file.readline():
            cur_line = line.rstrip()
            list_parameter_value = cur_line.split(':')
            dict_parameter_value[list_parameter_value[0]] = list_parameter_value[1]
        return dict_parameter_value


def write_in_end_row_file_txt(file_name = 'entry_ok.txt', list_with_data = []):
    """Записывает данные в конец строки текстового файла.
    Args:
        Имя файла, в который производить запись;
        Данные в ивде строки или в виде списка, которые необходимо записать."""
    with open(file_name, 'a', encoding='utf-8') as file:
        for index, data in enumerate(list_with_data):
            if isinstance(data, str):  # Если данные приходят в виде строки, 
                file.write(data)  # пишем строку в файл
            elif isinstance(data, list):  # Если данные в виде списка, 
                # то записываем предварительно преобразовав в строку
                file.write(str(data))
            # Если текущие данные последние в списке, 
            if index == len(list_with_data) - 1:
                # То переносим строку
                file.write('\n')
            else:  # Если текущие данные не последние в списке,
                # То разделяем двоеточием 
                file.write(':')


def read_xslx_with_filter_col1_col2(worksheet, num_col_filter, filter_value, num_col_target) -> list:
    """Возвращает список значений столбца, только тех строк, которые удовлетворяют условию.

    Args:
        worksheet - активный рабочий лист xslx, 
        num_col_filter - номер столбца, по которому фильтровать,
        filter_value - значение, по которому фильтровать (условие),
        num_col_target - номер целевого столбца.
    Return:
        list_with_filtered_values - список с отфильтрованными значениями.
    """

    if worksheet:  # Если активный лист получены
        current_row = 2  # С какой строки читать xlsx файл
        list_with_filtered_values.clear()

        # Пройтись со второй строки до первой пустой
        try: 
            # Находим ячейку в солбце 'Status'
            filter_cell = worksheet.cell(row=current_row, column=num_col_filter)
            while filter_cell.value is not None:
                try:
                    if filter_cell.value == filter_value:  # Если статус ок
                        target_cell = worksheet.cell(row=current_row, column=num_col_target)
                        list_with_filtered_values.append(target_cell.value)
                except:
                    logger.error(f'При проверке значений в {current_row} возникла ошибка')
                finally:
                    current_row += 1
                    filter_cell = worksheet.cell(row=current_row, column=num_col_filter)  # Переход на следующую строку
                    # pause = input('Нажмите любую клавишу для продолжения работы: ')
            else:
                logger.debug(f'Обнаружена пустая строка - {current_row} (конец xlsx файла).')
                # empty_row_is_finded = True
        except:
            logger.critical('КРИТИЧЕСКАЯ ОШИБКА: при выполнении фильтрации xlsx файла. Требуется перезапуск программы.')
    return list_with_filtered_values

import secrets
import string


def generate_random_password(number_symbols=12):
    """Генерирует случайный пароль из цифр и букв.
    Args:
    number_symbols : Int (число символов - длина пароля)
    Returns:
    random_password : str (сгенерированный пароль).
    """
    alphabet = string.ascii_letters + string.digits
    random_password = ''.join(secrets.choice(alphabet) for _ in range(number_symbols))
    return random_password

# import os 
# from dotenv import load_dotenv # Импортируем для безопасного хранения токенов
# load_dotenv()
# доступны в пространстве переменных окружения
# API_KEY_FOR_RUCAPTCHA = os.getenv('API_KEY_FOR_RUCAPTCHA')

from twocaptcha import TwoCaptcha  # для использования сервиса rucaptcha.ru
from time import sleep


def decrypt_captcha_deform_text(api_key_for_rucaptcha):
    """Отправляет изображение каптчи на rucaptcha для расшифровки.
    Args:
    api_key_for_rucaptcha : str - ключ для работы с API rucaptcha
    Returns:
    code : str - текстовая строка с расшифрованными символами."""
    solver = TwoCaptcha(api_key_for_rucaptcha)
    id = solver.send(file='image_captcha.jpg')
    sleep(12)
    code = solver.get_result(id)
    return code