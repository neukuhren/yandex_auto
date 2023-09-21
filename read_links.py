"""Модуль для чтения файла lnk.txt и переноса данных в lnk.xlsx файл"""


import logging # Импортируем библиотеку для безопасного хранения логов
import inspect

# from openpyxl import Workbook
from openpyxl import load_workbook
# from xlsxwriter import Workbook
# import xlrd

# from config import LIST_WITH_NAMES_OF_COLUMNS
from time import time


# В переменной __name__ хранится имя пакета;
# Это же имя будет присвоено логгеру.
# Это имя будет передаваться в логи, в аргумент %(name)
logger = logging.getLogger(__name__)

cur_row = 1


def read_txt_file_on_lines_and_return_list(input_file_name='lnk.txt') -> list:
    """Читает файл txt построчно и возвращает результат в виде списка.
    """
    try:
        with open(input_file_name) as file:
            list_with_links = []
            while line := file.readline():
                cur_line = line.rstrip()
                _link = cur_line  # Строка со сылкой                
                list_with_links.append(_link)
        return list_with_links  # Список со ссылками
    
    except:
        logger.critical('КРИТИЧЕСКАЯ ОШИБКА: при работе с исходным txt файлом. Экспорт данных невозможен.')
    

def write_to_excel(list_with_links, file_name='lnk.xlsx'):
    """Производит запись в xlsx файл"""

    # Найти целевой файл и рабочий лист
    try:
        workbook = load_workbook(file_name) 
        sheet_name = 'all'
        worksheet = workbook.active  # делаем единственный лист активным 
        if worksheet:
            logger.debug(f'Рабочий лист {sheet_name} найден')
        else:
            logger.debug(f'Рабочий лист {sheet_name} не найден')
    except:
        # logger.critical('КРИТИЧЕСКАЯ ОШИБКА: не найден выходной xlsx файл или лист. Экспорт данных невозможен.')
        pass

        # # Создание строки с заголовками
        # headers_list = list_with_names_of_columns
        # first_row = 0
        # for header in headers_list:
        #     col = headers_list.index(header) # Порядок (индекс) столбца
        #     worksheet.write(first_row, col, header) # Первая строка - заголовок

        # row = 1
        # for dict_ in list_with_data:
        #     for _key,_value in dict_.items():
        #         col=headers_list.index(_key)
        #         worksheet.write(row,col,_value)
        #     row += 1 # Перевод на следующую строку

    # Найти номер первой пустой строки
    try: 
        cur_row = 1
        cur_cell = worksheet.cell(row=cur_row, column=1)
        while cur_cell.value is not None:
            cur_row += 1
            cur_cell = worksheet.cell(row=cur_row, column=1)
        else:
            logger.info(f'Найден номер первой пустой строки - {cur_row} (для начала заполнения).')
            cur_row_is_finded = True
    except:
        logger.critical('КРИТИЧЕСКАЯ ОШИБКА: не найдена пустая строка для заполнения.')
        pass

    if cur_row_is_finded:
        print(f'ВНИМАНИЕ! Данные в файле {file_name} будут перезаписаны начиная со строки {cur_row}.')
        selected_start_empty_row = input(f'Введите y - для подтверждения, n - для отмены перезаписи. '\
                                         f'Или введите номер строки, с которой начать запись: ')
        
    # Если введено 'y' или номер строки
    if (selected_start_empty_row.lower() == 'y') or (int(selected_start_empty_row) > 0):
        if int(selected_start_empty_row) > 0:  
            cur_row = int(selected_start_empty_row) 
        # Итерируемся по списку заполняем строки таблицы
        try:
            for el_links in list_with_links:
                cur_cell = worksheet.cell(row=cur_row, column=1, value=el_links)
                cur_row += 1
            logger.INFO('Данные из файла txt успешно перенесены.')
        except:
            # logger.critical('КРИТИЧЕСКАЯ ОШИБКА: при записи данных в файл xlsx.')
            pass

        # worksheet['A10'].value = 123
        # cell_a10 = worksheet['A10']
        # print(f'в ячейке {cell_a10} найдено значение {cell_a10.value}')
        # d = worksheet.cell(row=10, column=3, value='12.05.2006')
        # print(d.value)
        try:
            workbook.save(file_name)
            logger.info(f'Книга {file_name} успешно сохранена.')
            workbook.close
        except:
            logger.critical('КРИТИЧЕСКАЯ ОШИБКА: при сохранении файла xlsx.')
    else:
        print(f'Запись данных в файл {file_name} отменена.')
    print('Работы программы по переносу данных завершена.')


def export_data_from_txt_file_in_xlsx_file(
        input_file_name='lnk.txt',
        output_file_name='lnk.xlsx'
        ):
    """Переносит данные из txt файла в файл xlsx"""

    list_with_links = read_txt_file_on_lines_and_return_list(input_file_name=input_file_name)
    """Список со ссылками."""
    # Открыть файл
    # В первую пустую строку вписать все данные из списка
    # Сохранить файл 
    write_to_excel(list_with_links=list_with_links, file_name=output_file_name)



def read_links_main():
    """Основная функция приложения read_lnk"""
    logger.debug(f'Запуск функции {inspect.currentframe().f_code.co_name}')
    export_data_from_txt_file_in_xlsx_file(
        input_file_name='lnk.txt',
        output_file_name='lnk.xlsx'
        )


if __name__ == '__main__':
    start_time = time()
    # Здесь задана глобальная конфигурация для всех логгеров
    logging.basicConfig(
        handlers=[logging.StreamHandler()],
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        # filename='main.log',
        # filemode='w'
        # w — содержимое файла перезаписывается при каждом запуске программы;
        # x — создать файл и записывать логи в него; если файл с таким именем уже существует — возникнет ошибка;
        # a — дописывать новые логи в конец указанного файла.
    )
    
    read_links_main()

    print('\n\n')
    print('Время работы программы составило: ')
    print("--- %s seconds ---" % round((time() - start_time), 0))
    print('[!] Программа выполнена.')
    